from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "trade_journal.csv"
XLSX_PATH = ROOT / "trade_journal.xlsx"

HEADERS = {
    "id": "编号",
    "entry_time_bjt": "开仓时间(北京时间)",
    "exit_time_bjt": "平仓时间(北京时间)",
    "symbol": "币种",
    "side": "方向",
    "quantity": "数量",
    "entry_price": "入场",
    "exit_price": "平仓",
    "stop_price": "止损",
    "result": "结果",
    "gross_pnl_usdt": "毛PnL(USDT)",
    "costs_usdt": "手续费/成本(USDT)",
    "net_pnl_usdt": "净PnL(USDT)",
    "account_impact": "账户影响",
    "trade_score": "评分",
    "core_conclusion": "核心结论",
    "success_failure_reason": "成功/失败原因",
    "next_avoidance": "下次规则",
}

WIDTHS = {
    "A": 10,
    "B": 22,
    "C": 22,
    "D": 14,
    "E": 10,
    "F": 10,
    "G": 12,
    "H": 12,
    "I": 12,
    "J": 18,
    "K": 14,
    "L": 16,
    "M": 14,
    "N": 34,
    "O": 10,
    "P": 54,
    "Q": 58,
    "R": 62,
}


def read_rows() -> list[dict[str, str]]:
    with CSV_PATH.open("r", newline="", encoding="utf-8") as file:
        return list(csv.DictReader(file))


def zh_value(field: str, value: str) -> str:
    maps = {
        "side": {
            "LONG": "做多",
            "SHORT": "做空",
        },
        "result": {
            "failed_but_risk_controlled": "失败但风控执行",
            "success": "成功",
            "failed": "失败",
            "breakeven": "保本",
            "LOSS": "失败",
            "LOSS_STOPPED": "失败止损",
            "PROFIT_PROTECTED_TOO_TIGHT": "盈利但保护过早",
        },
        "account_impact": {
            "about -0.63%": "约-0.63%",
            "small test loss; stop slipped below trigger": "小仓测试亏损且止损滑点",
            "small profit; protected winner": "小盈利但过早保护",
        },
    }
    return maps.get(field, {}).get(value, value)


def export() -> None:
    rows = read_rows()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "合约复盘"

    fields = list(HEADERS)
    sheet.append([HEADERS[field] for field in fields])
    for row in rows:
        sheet.append([zh_value(field, row.get(field, "")) for field in fields])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = PatternFill("solid", fgColor="F3F6FA")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, sheet.max_row + 1):
        if row_idx % 2 == 0:
            for cell in sheet[row_idx]:
                cell.fill = thin_gray
        for cell in sheet[row_idx]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col, width in WIDTHS.items():
        sheet.column_dimensions[col].width = width

    for row in sheet.iter_rows(min_row=2):
        row[4].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        row[14].alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28
    for row_idx in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].height = 86

    summary = workbook.create_sheet("统计")
    summary.append(["指标", "值"])
    summary.append(["交易笔数", len(rows)])
    wins = sum(1 for row in rows if float(row.get("net_pnl_usdt") or 0) > 0)
    losses = sum(1 for row in rows if float(row.get("net_pnl_usdt") or 0) < 0)
    net_total = sum(float(row.get("net_pnl_usdt") or 0) for row in rows)
    summary.append(["盈利笔数", wins])
    summary.append(["亏损笔数", losses])
    summary.append(["胜率", f"{wins / len(rows) * 100:.1f}%" if rows else "0.0%"])
    summary.append(["累计净PnL(USDT)", round(net_total, 6)])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 18

    workbook.save(XLSX_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    export()
